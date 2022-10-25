#常用正则表达式
#单字符：
#       . :除换行以外所有字符
#       []:[aoe] [a-w]匹配集合中任意一个字符
#       \d:数字[0-9]
#       \D:非数字
#       \w:数字、字母、下划线、中文
#       \W:非\w
#       \s:所有的空白字符包括空格、制表符、换页符等等，等价于[ \f\n\r\t\v]
#       \S:非空白
#数量修饰
#       *    :任意多次 >=0
#       +    :至少一次 >=1
#       ?    :可有可无 0次或者1次
#       {m}  :固定m次
#       {m,} :至少m次
#       {m,n}:m-n次
#边界
#       $  :以某某结尾
#       ^  :以某某开头
#分组
#       (ab)
#贪婪模式:  .*
#非贪婪（惰性）模式:  .*?

#re.I : 忽略大小写
#re.M : 多行匹配
#re.S : 单行匹配
#re.sub(正则表达式，替换内容，字符串)

import requests
import re
import openpyxl
import os

#爬指定文字段
def txt():
    iter1 = re.compile(r'<li>.*?<span class="title">(?P<movie_name>.*?)</span>.*?'
                       r'<br>(?P<movie_year>.*?)&nbsp;.*?'
                       r'<span class="rating_num" property="v:average">(?P<movie_rate>.*?)</span>.*?'
                       r'<span>(?P<movie_num>.*?)</span>' , re.S) #逐行往下匹配
    wb=openpyxl.load_workbook('./txt.xlsx')
    sh=wb['Sheet1']
    sh.cell(row=1,column=1).value='电影名'
    sh.cell(row=1,column=2).value='上映时间'
    sh.cell(row=1, column=3).value = '评分'
    sh.cell(row=1, column=4).value = '观看数量'
    x=2
    for start in range(0 , 250,25):
        url1 = 'https://movie.douban.com/top250?'
        param1 = {
            'start': start,
            'filter': ''
        }
        header = {
            'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36 Edg/105.0.1343.27'
        }
        resp = requests.get(url = url1 ,params=param1,headers = header)
        page_text = resp.text
        resp.close()
        result = iter1.finditer(page_text)
        for item in result:
            movie_detail = item.groupdict('')
            movie_detail['movie_year'] = movie_detail['movie_year'].strip()
            sh.cell(row=x,column=1).value=movie_detail['movie_name']
            sh.cell(row=x, column=2).value = movie_detail['movie_year']
            sh.cell(row=x, column=3).value = movie_detail['movie_rate']
            sh.cell(row=x, column=4).value = movie_detail['movie_num']
            print('第%d部电影信息已爬取'%(x-1))
            x+=1
            #csv_handle.writerow(movie_detail.values())
    wb.save('./txt.xlsx')
    print('Over')

#爬图片
#re.search()	在一个字符串中搜索匹配正则表达式的第一个位置，返回match对象
#re.match()	    从一个字符串的开始位置起匹配正则表达式，返回match对象
#re.findall()	搜索字符串，以列表类型返回全部能匹配的子串,    常用
#re.split()	    将一个字符串按照正则表达式匹配结果进行分割，返回列表类型
#re.finditer()	搜索字符串，返回一个匹配结果的迭代类型，每个迭代元素式match对象
#re.sub()	    在一个字符串中替换所有匹配正则表达式的子串，返回替换后的字符串
#re.compile(pattern, flags=0)   flags匹配模式，如re.S
def paper1():
    #创建一个文件夹存储图片
    if not os.path.exists('./qiutu'):
        os.mkdir('./qiutu')
    headers={
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36 Edg/105.0.1343.27'
    }
    #通用url助力翻页
    url='https://www.qiushibaike.com/pic/page/%d/?s=5184961'
    #35页数据
    for pageNum in range(1,36):
        new_url=format(url%pageNum)
        #把图片所在界面代码爬下来
        page_text=requests.get(url=new_url,headers=headers).text
        #使用正则化规则进行图片的网址爬下来,括号里面是我们想要的东西
        ex='<div class="thumb">.*?<img src="(.*?)" alt.*?</div>'
        img_src_list=re.findall(ex,page_text,re.S)
        #逐个读取列表中的图片地址并下载
        for src in img_src_list:
            #拼接出一个完整url
            src='https:'+src
            img_data=requests.get(url=src,headers=headers).content
            #生成图片名称并存取
            img_name=src.split('/')[-1]
            imgPath='./qiutu/'+img_name
            with open(imgPath,'wb') as fp:
                fp.write(img_data)
                print(img_name,'下载成功！！')
#/P<group_name> : 给要爬取的部分定义一个标签group_name
def paper2_1():
    url1 = 'https://movie.douban.com/top250?'
    #ua封装
    #对于jpg or png，用requests.get().content来返回图片的二进制式图片数据，text返回字符串形式数据，json返回对象类型数据
    header1 = {
        'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36 QIHU 360SE/13.1.6140.0'
    }
    obj = re.compile(r'<li>.*?<img width="100" alt=".*?" src="(?P<movie_img>.*?)" class="">.*?'
                     r'<span class="title">(?P<movie_name>.*?)</span>.*?</li>', re.S) #re.S逐行匹配
    for start in range(0,250,25): #起点、终点、步进，注意区间左闭右开，所以取值其实是0 25 50 75 …… 没有250！
        param1={
            'start': start,
            'filter':''
        }
        #解析数据
        resp = requests.get(url = url1, params=param1, headers = header1)
        page_text = resp.text
        resp.close()
        #开始匹配
        result = obj.finditer(page_text)
        for it in result:
            url2=it.group('movie_img')#获得url
            print(url2)
            name=it.group('movie_name')
            print(name)
            img_data=requests.get(url=url2, headers=header1).content
            with open('./tupian/%s.jpg'%name,'wb')as fp:
                fp.write(img_data)
            print('%s封面图下载完成'%name)
    print('over')

def paper2_2():
    url1 = 'https://movie.douban.com/top250?'
    header1 = {
       'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36 QIHU 360SE/13.1.6140.0'
    }
    #图片地址存储
    ex='<li>.*?<img width="100" alt=".*?" src="(.*?)" class="">.*?</li>'
    #名字存储
    name='<li>.*?<span class="title">(.*?)</span>.*?</li>'
    for i in range(0,250,25):
        param1={
            'start': i ,
            'filter':''
            }
        #解析数据
        resp = requests.get(url = url1, params=param1, headers = header1)
        page_text = resp.text
        resp.close()
        #开始匹配
        img_src_list=re.findall(ex,page_text,re.S)
        img_name_list=re.findall(name,page_text,re.S)
        print(img_src_list)
        print(img_name_list)
        for i in range(len(img_src_list)):
            img_data=requests.get(url=img_src_list[i], headers=header1).content
            with open('./tupian/%s.jpg'%img_name_list[i],'wb')as fp:
                fp.write(img_data)
            print('%s封面图片下载完成'%img_name_list[i])
    print('over')
