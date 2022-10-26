import requests
import asyncio
from lxml import etree
import openpyxl
from Big_homework.Login_cookie import Login_cookie

file_path='data.xlsx'
wb=openpyxl.load_workbook(file_path)
ws=wb['Sheet1']

name='13326786127'
password='Xiejiajian34'
cookies,page,browser=Login_cookie(name,password).get_cookie()
headers={
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36'
}


async def main(down_times=0):
    await asyncio.sleep(5)
    await page.waitForSelector('#Popover1-toggle',{'timeout':3000})
    await page.click('#Popover1-toggle')
    await page.type('#Popover1-toggle','如何看待三胎政策')
    await page.click('#root > div > div:nth-child(2) > header > div.AppHeader-inner.css-l2ygoj > div.css-1acwmmj > div > form > div > div > label > button')
    #下滑
    for i in range(down_times):
        await asyncio.sleep(6)
        await page.evaluate("""{window.scrollBy(0, document.body.scrollHeight);}""")
    #获得目前显示的话题名单
    await asyncio.sleep(10)
    continue_get=True
    i=1#从第几个评论开始爬
    n=1#总共爬多少个记录
    # 页面跳转
    while continue_get:
        try:
            print(f'第{i}个评论开始爬取')
            click_handle=await page.xpath(f'//*[@id="SearchMain"]/div/div/div/div/div[{i}]/div/div/div/h2/span/div/a/span')
            await click_handle[0].click()
            print('clikc')
            await asyncio.sleep(3)
            #页面跳转
            page_list=await browser.pages()
            print(page_list)
            page2=page_list[2]
            await asyncio.sleep(2)
            #评论获取
            #查看所有回答
            try:
                click1= await page2.xpath(
                    '//*[@id="root"]/div/main/div/div/div[3]/div[1]/div/div[1]/a')
                await click1[0].click()
                print('查看所有回答')
                await asyncio.sleep(3)
                #下拉
                while True:
                    try:
                        await page2.evaluate("""{window.scrollBy(0, document.body.scrollHeight);}""")
                        print('下拉')
                        await asyncio.sleep(3)
                    except:
                        print('已滑到最底')
                        break
                num = 1  # 记录评论数
                cont = True
                while cont:
                    try:
                        a = await page2.xpath(
                            f'// *[ @ id = "QuestionAnswers-answers"] / div / div / div / div[2] / div / div[{num}] / div / div / div[1] / \
                                                                          div[1] / div[1] / div / div[1] / span / div / a')
                        name = await (await a[0].getProperty('textContent')).jsonValue()
                        print(name)

                        b = await page2.xpath(
                            f'//*[@id="QuestionAnswers-answers"]/div/div/div/div[2]/div/div[{num}]/div/div/div[2]/div[1]/div/a/span')
                        time = await (await b[0].getProperty('textContent')).jsonValue()
                        print(time)

                        c = await page2.xpath(
                            f'// *[ @ id = "QuestionAnswers-answers"] / div / div / div / div[2] / div / div[{num}] / div / div / div[2] / div[2] / span / button')
                        like = await (await c[0].getProperty('textContent')).jsonValue()
                        print(like)

                        d = await page2.xpath(
                            '// *[ @ id = "QuestionAnswers-answers"] / div / div / div / div[2] / div / div[{num}] / div / div / div[2] / span[1] / div / span / p')
                        content = ""
                        for item in d:
                            text = await (await item.getProperty('textContent')).jsonValue()
                            print(text)
                            content += text
                        print(content)

                        ws.cell(row=n + 1,column=1).value  = name
                        ws.cell(row=n + 1, column=2).value = time
                        ws.cell(row=n + 1, column=3).value = content
                        ws.cell(row=n + 1, column=4).value = like

                        num += 1
                        n+=1

                    except:
                        print('已爬取全部评论文章')
                        cont = False
            except:
                print('不可以查看所有回答')


            try:
                # 获得评论
                # content=await page2.content()
                await asyncio.sleep(2)
                a = await page2.xpath('//*[@id="root"]/div/main/div/article/header/div[1]/div/div/div/div[1]/span/div')
                name = await (await a[0].getProperty('textContent')).jsonValue()
                print(name)

                b = await page2.xpath('//*[@id="root"]/div/main/div/article/div')
                time = await (await b[1].getProperty('textContent')).jsonValue()
                print(time)

                c = await page2.xpath('//*[@id="root"]/div/main/div/article/div[4]/div/div/span/button')
                like = await (await c[0].getProperty('textContent')).jsonValue()
                print(like)

                d = await page2.xpath('//*[@id="root"]/div/main/div/article/div[1]/div/div/p')
                content = ""
                for item in d:
                    text = await (await item.getProperty('textContent')).jsonValue()
                    print(text)
                    content += text
                print(content)

                ws.cell(row=n + 1, column=1).value = name
                ws.cell(row=n + 1, column=2).value = time
                ws.cell(row=n + 1, column=3).value = content
                ws.cell(row=n + 1, column=4).value = like

                n+=1

            except:
                print("error")


            #关闭页面
            await page2.close()
            i+=1
            await asyncio.sleep(2)
            #if i%3==0:

        except Exception as e:
            print('评论迭代完毕')
            continue_get=False
            print(e)

    wb.save(file_path)




asyncio.get_event_loop().run_until_complete(main())